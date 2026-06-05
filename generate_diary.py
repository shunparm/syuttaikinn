#!/usr/bin/env python3
"""技能実習日誌 自動生成スクリプト

「月次入力」シートの出勤情報から番号・業務名・指導内容を自動算出し、
★日誌自動生成のSTEP4振り分け表に書き込む。
--transfer オプションで対象月の日誌シートへも自動転記。

使い方:
    python generate_diary.py <Excelファイル>
    python generate_diary.py <Excelファイル> --transfer     # 日誌シートへ転記
    python generate_diary.py <Excelファイル> --seed 42
"""

import calendar
import math
import random
import sys

import openpyxl

INPUT_SHEET = "月次入力"    # 年・月・出勤日を入力するシート
AUTO_SHEET  = "★日誌自動生成"  # STEP4 出力先シート

MONTHLY_REQUIRED = {
    1:  {1: 91, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    2:  {1: 95, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    3:  {1: 95, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    4:  {1: 95, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    5:  {1: 94, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    6:  {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    7:  {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    8:  {1: 70, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    9:  {1: 70, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    10: {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    11: {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    12: {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
}

NUM_TO_WORKS = {
    1: [
        ("掘削作業", ["マンホール布設", "汚水桝布設", "管布設", "水道掘削", "溝掘削", "法面掘削", "基礎掘削"]),
        ("土砂積込み作業", ["過積載防止", "周囲の確認", "積込み確認", "安全確認"]),
        ("走行操作作業", ["発進操作", "平坦地走行", "登坂操作", "降坂操作", "停止操作", "下車操作"]),
        ("毎日整備", ["目視点検", "グリース注入", "燃料補給", "清掃作業"]),
        ("始業前点検", ["目視点検", "備品確認", "始業確認", "安全点検"]),
    ],
    2: [
        ("作業開始前の安全装置等の点検作業", ["目視点検", "安全確認", "始業前確認", "安全装置確認"]),
        ("建設機械施工職種に必要な整理整頓作業", ["道具整理", "現場整理", "用具整頓", "工具整備"]),
        ("保護具の着用と服装の安全点検作業", ["保護具確認", "服装点検", "安全確認"]),
        ("雇入れ時等の安全衛生教育", ["安全教育", "衛生教育", "ルール確認"]),
    ],
    3: [
        ("掘削作業", ["手元作業", "補助作業", "埋戻し", "残土処理"]),
        ("締固め作業", ["埋戻し", "ランマ転圧", "転圧作業", "締固め確認"]),
        ("土工作業(対象職種・作業に係る手作業の作業）", ["手元作業", "蓋かさ上げ", "補助作業", "整地作業"]),
        ("積込み作業", ["手元作業", "補助積込み", "残土積込み"]),
        ("建設機械の管理及び点検・整備作業", ["バケット交換", "グリース注入", "清掃作業", "オイル点検"]),
    ],
    4: [
        ("安全衛生業務", ["荷物搬入", "現場清掃", "補助作業", "用具整備", "資材確認"]),
        ("建設機械施工職種に必要な整理整頓作業", ["道具整理", "現場整理", "工具整備"]),
    ],
    5: [
        ("建設機械の移送車両への積載及び移送作業", ["声出し誘導", "ユンボ回送", "機械移送", "積載補助", "誘導補助"]),
    ],
    6: [
        ("安全衛生業務", ["安全訓練", "倉庫整理", "現場内清掃", "安全管理", "安全確認"]),
    ],
}

def allocate_days(n_days: int, month: int) -> dict:
    reqs = MONTHLY_REQUIRED[month]
    min_d = {k: math.ceil(v / 8) for k, v in reqs.items()}
    total_min = sum(min_d.values())

    if n_days >= total_min:
        alloc = dict(min_d)
        alloc[1] += n_days - total_min
    elif n_days > 0:
        alloc = dict(min_d)
        deficit = total_min - n_days
        take_from_1 = min(deficit, alloc[1] - 1)
        alloc[1] -= take_from_1
        deficit -= take_from_1
        if deficit > 0:
            take_from_3 = min(deficit, alloc[3] - 1)
            alloc[3] -= take_from_3
    else:
        alloc = {k: 0 for k in reqs}

    return alloc


def shuffle_no_triple(lst: list, rng: random.Random) -> list:
    for _ in range(200):
        rng.shuffle(lst)
        if all(
            not (i >= 2 and lst[i] == lst[i - 1] == lst[i - 2])
            for i in range(len(lst))
        ):
            return lst
    return lst


def read_working_days(ws, days_in_month: int) -> list:
    """月次入力シートの行9/10（1〜15日）と行11/12（16〜31日）から出勤日リストを返す。

    行9/11に日付番号、行10/12に出勤フラグ（1=出勤）が入る構造。
    列は1始まり（A=1）。
    """
    working = []
    for col in range(1, 16):
        day = ws.cell(row=9, column=col).value
        flag = ws.cell(row=10, column=col).value
        if day and (flag == 1 or flag == "1") and int(day) <= days_in_month:
            working.append(int(day))
    for col in range(1, 17):
        day = ws.cell(row=11, column=col).value
        flag = ws.cell(row=12, column=col).value
        if day and (flag == 1 or flag == "1") and int(day) <= days_in_month:
            working.append(int(day))
    return sorted(working)


def write_step4(ws, year: int, month: int, working_days: list,
                alloc: dict, num_seq: list, work_idx: dict):
    """★日誌自動生成のSTEP4振り分け表（行27〜57）に結果を書き込む。"""
    days_in_month = calendar.monthrange(year, month)[1]
    OUTPUT_ROW = 27

    for day in range(1, 32):
        row = OUTPUT_ROW + day - 1

        if day > days_in_month:
            ws.cell(row=row, column=4).value = None
            ws.cell(row=row, column=10).value = None
            ws.cell(row=row, column=11).value = None
            continue

        # B列(曜日)・C列(出/休)はExcelの数式で自動計算されるため書き込まない

        if day in working_days:
            seq_idx = working_days.index(day)
            num = num_seq[seq_idx]
            works = NUM_TO_WORKS[num]
            wi = work_idx[num] % len(works)
            wname, shidou_list = works[wi]
            shidou = shidou_list[work_idx[num] % len(shidou_list)]
            work_idx[num] += 1

            ws.cell(row=row, column=4).value = num
            ws.cell(row=row, column=10).value = wname
            ws.cell(row=row, column=11).value = shidou
        else:
            ws.cell(row=row, column=4).value = None
            ws.cell(row=row, column=10).value = None
            ws.cell(row=row, column=11).value = None


def transfer_to_diary(wb, ws_auto, year: int, month: int, supervisor: str,
                      working_days: list):
    """★日誌自動生成のSTEP4データを対象月の日誌シートへ転記する。

    出勤判定は月次入力シートから取得した working_days を使う
    （STEP4 C列はExcel数式のためopenpyxlでは評価値を読めないため）。
    """
    candidates = [
        f"{year}.{month}",
        f"{year}.{month} ",
    ]
    ws_diary = None
    found_name = None
    for name in candidates:
        if name in wb.sheetnames:
            ws_diary = wb[name]
            found_name = name
            break

    if ws_diary is None:
        print(f"警告: 日誌シート {year}.{month} が見つかりません。転記をスキップします。")
        print(f"  探索した名前: {candidates}")
        return

    days_in_month = calendar.monthrange(year, month)[1]
    OUTPUT_ROW = 27
    DIARY_ROW = 9

    transferred = 0
    for day in range(1, days_in_month + 1):
        src_row = OUTPUT_ROW + day - 1
        dst_row = DIARY_ROW + day - 1

        # C列は数式のためopenpyxlで読めない → 月次入力の working_days で出勤判定
        num    = ws_auto.cell(row=src_row, column=4).value
        wname  = ws_auto.cell(row=src_row, column=10).value
        shidou = ws_auto.cell(row=src_row, column=11).value

        if day in working_days and num is not None:
            ws_diary.cell(row=dst_row, column=2).value = wname
            ws_diary.cell(row=dst_row, column=4).value = num
            ws_diary.cell(row=dst_row, column=5).value = shidou
            ws_diary.cell(row=dst_row, column=8).value = supervisor
            transferred += 1
        else:
            ws_diary.cell(row=dst_row, column=2).value = "休み"
            ws_diary.cell(row=dst_row, column=4).value = None
            ws_diary.cell(row=dst_row, column=5).value = None
            ws_diary.cell(row=dst_row, column=8).value = None

    print(f"転記完了: シート「{found_name}」 ({transferred}日分)")


def print_summary(month: int, working_days: list, alloc: dict):
    reqs = MONTHLY_REQUIRED[month]
    print("\n【番号別 集計】")
    print(f"{'番号':>4}  {'必要時間':>8}  {'割当日数':>8}  {'実績時間':>8}  {'達成':>4}")
    print("-" * 46)
    for num in range(1, 7):
        req_h = reqs[num]
        assigned = alloc.get(num, 0)
        actual_h = assigned * 8
        ok = "✓" if actual_h >= req_h else "△"
        print(f"{num:>4}  {req_h:>7}h  {assigned:>7}日  {actual_h:>7}h  {ok:>4}")
    print("-" * 46)
    total_req = sum(reqs.values())
    total_actual = sum(alloc.get(n, 0) * 8 for n in range(1, 7))
    all_ok = "✓" if all(alloc.get(n, 0) * 8 >= reqs[n] for n in range(1, 7)) else "△"
    print(f"{'合計':>4}  {total_req:>7}h  {len(working_days):>7}日  {total_actual:>7}h  {all_ok:>4}")


def parse_args():
    args = list(sys.argv[1:])
    seed = None
    do_transfer = False

    if "--seed" in args:
        idx = args.index("--seed")
        seed = int(args[idx + 1])
        args.pop(idx + 1)
        args.pop(idx)
    if "--transfer" in args:
        do_transfer = True
        args.remove("--transfer")

    if len(args) < 1:
        print("使い方: python generate_diary.py <Excelファイル> [--transfer] [--seed N]")
        sys.exit(1)

    return args[0], seed, do_transfer


def main():
    excel_path, seed, do_transfer = parse_args()
    rng = random.Random(seed)

    wb = openpyxl.load_workbook(excel_path)

    for sheet_name in (INPUT_SHEET, AUTO_SHEET):
        if sheet_name not in wb.sheetnames:
            print(f"エラー: シート「{sheet_name}」が見つかりません。")
            sys.exit(1)

    ws_input = wb[INPUT_SHEET]   # 読み取り専用（年・月・出勤日）
    ws_auto  = wb[AUTO_SHEET]    # 書き込み先（STEP4）

    year       = int(ws_input["B5"].value)
    month      = int(ws_input["E5"].value)
    supervisor = ws_input["I5"].value or "中原"
    days_in_month = calendar.monthrange(year, month)[1]

    working_days = read_working_days(ws_input, days_in_month)
    n_working = len(working_days)

    print(f"対象: {year}年{month}月 / 指導員: {supervisor}")
    print(f"出勤日数: {n_working}日 / 出勤日: {working_days}")

    alloc = allocate_days(n_working, month)

    num_seq = []
    for num, count in alloc.items():
        num_seq.extend([num] * count)
    num_seq = shuffle_no_triple(num_seq, rng)

    work_idx = {k: 0 for k in NUM_TO_WORKS}

    write_step4(ws_auto, year, month, working_days, alloc, num_seq, work_idx)

    if do_transfer:
        transfer_to_diary(wb, ws_auto, year, month, supervisor, working_days)

    wb.save(excel_path)
    print(f"\n書き込み完了: {excel_path}")
    print_summary(month, working_days, alloc)


if __name__ == "__main__":
    main()
