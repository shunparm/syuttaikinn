#!/usr/bin/env python3
"""技能実習日誌 新月シート追加スクリプト

既存の日誌シートをテンプレートとしてコピーし、
番号割り当て・業務名・指導内容を自動入力して新月シートを追加する。

実習生4人で同内容のため、月1枚シートを作成し氏名を変えて4回印刷する運用。

使い方:
    python add_month.py <Excelファイル> <年> <月> <氏名> [出勤日] [--seed N]

例:
    python add_month.py 技能実習日誌.xlsx 2026 5 ヨザ 1,2,5,6,7,8,9,12,13,14,15,16,19,20,21,22,23
    python add_month.py 技能実習日誌.xlsx 2026 5 ヨザ --seed 42 1,2,5,6,7

氏名だけ変更して再印刷する場合:
    python add_month.py 技能実習日誌.xlsx 2026 5 リズキ --rename-only
"""

import calendar
import datetime
import math
import random
import sys

import openpyxl

SUPERVISOR = "中原"

# ★・■で始まるシートはテンプレート対象外
EXCLUDED_SHEETS = {"★日誌自動生成", "■入力マスタ"}

MONTHLY_REQUIRED = {
    1:  {1: 91, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    2:  {1: 95, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    3:  {1: 95, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    4:  {1: 95, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    5:  {1: 94, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    6:  {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    7:  {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    8:  {1: 70, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    9:  {1: 70, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    10: {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    11: {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    12: {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
}

NUM_TO_WORKS = {
    1: [
        ("掘削作業", ["マンホール布設", "汚水桝布設", "管布設", "水道掘削", "溝掘削", "法面掘削", "基礎掘削"]),
        ("土砂積込み作業", ["過積載防止", "周囲の確認", "積込み確認", "安全確認"]),
        ("走行操作作業", ["発進操作", "平坦地走行", "登坂操作", "降坂操作", "停止操作", "下車操作"]),
        ("毎日整備", ["目視点検", "グリース注入", "燃料補給", "清掃作業"]),
        ("始業前点検", ["目視点検", "備品確認", "始業確認", "安全点検"]),
    ],
    2: [
        ("作業開始前の安全装置等の点検作業", ["目視点検", "安全確認", "始業前確認", "安全装置確認"]),
        ("建設機械施工職種に必要な整理整頓作業", ["道具整理", "現場整理", "用具整頓", "工具整備"]),
        ("保護具の着用と服装の安全点検作業", ["保護具確認", "服装点検", "安全確認"]),
        ("雇入れ時等の安全衛生教育", ["安全教育", "衛生教育", "ルール確認"]),
    ],
    3: [
        ("掘削作業", ["手元作業", "補助作業", "埋戻し", "残土処理"]),
        ("締固め作業", ["埋戻し", "ランマ転圧", "転圧作業", "締固め確認"]),
        ("土工作業(対象職種・作業に係る手作業の作業）", ["手元作業", "蓋かさ上げ", "補助作業", "整地作業"]),
        ("積込み作業", ["手元作業", "補助積込み", "残土積込み"]),
        ("建設機械の管理及び点検・整備作業", ["バケット交換", "グリース注入", "清掃作業", "オイル点検"]),
    ],
    4: [
        ("安全衛生業務", ["荷物搬入", "現場清掃", "補助作業", "用具整備", "資材確認"]),
        ("建設機械施工職種に必要な整理整頓作業", ["道具整理", "現場整理", "工具整備"]),
    ],
    5: [
        ("建設機械の移送車両への積載及び移送作業", ["声出し誘導", "ユンボ回送", "機械移送", "積載補助", "誘導補助"]),
    ],
    6: [
        ("安全衛生業務", ["安全訓練", "倉庫整理", "現場内清掃", "安全管理", "安全確認"]),
    ],
}


def allocate_days(n_days: int, month: int) -> dict:
    reqs = MONTHLY_REQUIRED[month]
    min_d = {k: math.ceil(v / 8) for k, v in reqs.items()}
    total_min = sum(min_d.values())

    if n_days >= total_min:
        alloc = dict(min_d)
        alloc[1] += n_days - total_min
    elif n_days > 0:
        alloc = dict(min_d)
        deficit = total_min - n_days
        take_from_1 = min(deficit, alloc[1] - 1)
        alloc[1] -= take_from_1
        deficit -= take_from_1
        if deficit > 0:
            take_from_3 = min(deficit, alloc[3] - 1)
            alloc[3] -= take_from_3
    else:
        alloc = {k: 0 for k in reqs}

    return alloc


def shuffle_no_triple(lst: list, rng: random.Random) -> list:
    for _ in range(200):
        rng.shuffle(lst)
        if all(
            not (i >= 2 and lst[i] == lst[i - 1] == lst[i - 2])
            for i in range(len(lst))
        ):
            return lst
    return lst


def find_template(wb) -> openpyxl.worksheet.worksheet.Worksheet:
    """テンプレートとなる既存の日誌シートを返す（最新月シートを優先）。"""
    candidates = [
        ws for ws in wb.worksheets
        if ws.title not in EXCLUDED_SHEETS
        and not ws.title.startswith("★")
        and not ws.title.startswith("■")
    ]
    if not candidates:
        raise ValueError("テンプレートになる日誌シートが見つかりません。")
    return candidates[0]


def make_sheet_name(year: int, month: int) -> str:
    return f"{year}.{month}"


def update_header(ws, year: int, month: int, trainee_name: str):
    """ヘッダー部分（年月・氏名）だけ書き換える。"""
    ws["D5"] = f"（　　　{year}　　年　　{month}　　月分）"
    ws["E6"] = f"　　　　　　氏名　{trainee_name}"


def reset_data_rows(ws, year: int, month: int, working_days: list, rng: random.Random):
    """行9〜39のA〜H列を再設定する。I・J列（参照リスト）は触らない。"""
    days_in_month = calendar.monthrange(year, month)[1]
    alloc = allocate_days(len(working_days), month)

    num_seq = []
    for num, count in alloc.items():
        num_seq.extend([num] * count)
    num_seq = shuffle_no_triple(num_seq, rng)

    work_idx = {k: 0 for k in NUM_TO_WORKS}

    for day in range(1, 32):
        row = 8 + day
        if day > days_in_month:
            ws.cell(row=row, column=1).value = None
            ws.cell(row=row, column=2).value = None
            ws.cell(row=row, column=4).value = None
            ws.cell(row=row, column=5).value = None
            ws.cell(row=row, column=8).value = None
            continue

        ws.cell(row=row, column=1).value = datetime.datetime(year, month, day)

        if day in working_days:
            seq_idx = working_days.index(day)
            num = num_seq[seq_idx]

            works = NUM_TO_WORKS[num]
            wi = work_idx[num] % len(works)
            wname, shidou_list = works[wi]
            shidou = shidou_list[work_idx[num] % len(shidou_list)]
            work_idx[num] += 1

            ws.cell(row=row, column=2).value = wname
            ws.cell(row=row, column=4).value = num
            ws.cell(row=row, column=5).value = shidou
            ws.cell(row=row, column=8).value = SUPERVISOR
        else:
            ws.cell(row=row, column=2).value = "休み"
            ws.cell(row=row, column=4).value = None
            ws.cell(row=row, column=5).value = None
            ws.cell(row=row, column=8).value = None

    return alloc


def print_summary(month: int, working_days: list, alloc: dict):
    reqs = MONTHLY_REQUIRED[month]
    print("\n【番号別 集計】")
    print(f"{'番号':>4}  {'必要時間':>8}  {'割当日数':>8}  {'実績時間':>8}  {'達成':>4}")
    print("-" * 46)
    for num in range(1, 7):
        req_h = reqs[num]
        assigned = alloc.get(num, 0)
        actual_h = assigned * 8
        ok = "✓" if actual_h >= req_h else "△"
        print(f"{num:>4}  {req_h:>7}h  {assigned:>7}日  {actual_h:>7}h  {ok:>4}")
    print("-" * 46)
    total_req = sum(reqs.values())
    total_actual = sum(alloc.get(n, 0) * 8 for n in range(1, 7))
    all_ok = "✓" if all(alloc.get(n, 0) * 8 >= reqs[n] for n in range(1, 7)) else "△"
    print(f"{'合計':>4}  {total_req:>7}h  {len(working_days):>7}日  {total_actual:>7}h  {all_ok:>4}")


def parse_args():
    args = list(sys.argv[1:])
    seed = None
    rename_only = False

    if "--seed" in args:
        idx = args.index("--seed")
        seed = int(args[idx + 1])
        args.pop(idx + 1)
        args.pop(idx)

    if "--rename-only" in args:
        rename_only = True
        args.remove("--rename-only")

    if len(args) < 4:
        print("使い方: python add_month.py <Excelファイル> <年> <月> <氏名> [出勤日] [--seed N]")
        print("        python add_month.py <Excelファイル> <年> <月> <氏名> --rename-only  # 氏名だけ変更")
        print("例:     python add_month.py 技能実習日誌.xlsx 2026 5 ヨザ 1,2,5,6,7,8,9,12,13")
        sys.exit(1)

    excel_path = args[0]
    year = int(args[1])
    month = int(args[2])
    trainee_name = args[3]
    working_days_str = args[4] if len(args) >= 5 else None

    return excel_path, year, month, trainee_name, working_days_str, seed, rename_only


def main():
    excel_path, year, month, trainee_name, working_days_str, seed, rename_only = parse_args()
    rng = random.Random(seed)

    wb = openpyxl.load_workbook(excel_path)
    sheet_name = make_sheet_name(year, month)

    # --rename-only: 既存シートの氏名だけ変更して終了
    if rename_only:
        if sheet_name not in wb.sheetnames:
            print(f"エラー: シート「{sheet_name}」が見つかりません。先にシートを作成してください。")
            sys.exit(1)
        ws = wb[sheet_name]
        update_header(ws, year, month, trainee_name)
        wb.save(excel_path)
        print(f"氏名を「{trainee_name}」に変更しました → {excel_path}")
        return

    # 出勤日の取得
    if working_days_str:
        working_days = sorted(int(d.strip()) for d in working_days_str.split(",") if d.strip())
    else:
        days_in_month = calendar.monthrange(year, month)[1]
        print(f"{year}年{month}月（全{days_in_month}日）")
        raw = input("出勤日を入力してください（例: 1,2,5,6,7,...）: ")
        working_days = sorted(int(d.strip()) for d in raw.split(",") if d.strip())

    days_in_month = calendar.monthrange(year, month)[1]
    invalid = [d for d in working_days if d < 1 or d > days_in_month]
    if invalid:
        print(f"エラー: {month}月に存在しない日付 {invalid}")
        sys.exit(1)

    if sheet_name in wb.sheetnames:
        print(f"警告: シート「{sheet_name}」は既に存在します。上書きします。")
        del wb[sheet_name]

    # テンプレートをコピー（★・■シートを除いた最新日誌シート）
    template_ws = find_template(wb)
    print(f"テンプレート: 「{template_ws.title}」をコピー")
    new_ws = wb.copy_worksheet(template_ws)
    new_ws.title = sheet_name

    # 先頭に移動
    wb.move_sheet(new_ws, offset=-len(wb.sheetnames) + 1)

    update_header(new_ws, year, month, trainee_name)
    alloc = reset_data_rows(new_ws, year, month, working_days, rng)

    wb.save(excel_path)

    print(f"\n対象: {year}年{month}月 / 氏名: {trainee_name}")
    print(f"出勤日数: {len(working_days)}日 / 出勤日: {working_days}")
    print(f"シート「{sheet_name}」を先頭に追加しました → {excel_path}")
    print_summary(month, working_days, alloc)


if __name__ == "__main__":
    main()
