#!/usr/bin/env python3
"""
技能実習日誌 一括生成スクリプト（出勤簿CSV版）

出勤簿アプリからエクスポートしたCSVを読み込み、
実習生全員分の技能実習日誌シートを一括で生成する。

使い方:
  python run.py 出勤簿.csv
  python run.py 出勤簿.csv --trainees EMP008,EMP009,EMP010,EMP011
  python run.py 出勤簿.csv --supervisor 中原
  python run.py 出勤簿.csv --excel 日誌テンプレート.xlsx
  python run.py 出勤簿.csv --seed 42

出力: Excelファイルに「{年}.{月}_{名前}」シートを追加
      同名シートが既に存在する場合は上書き
"""

import argparse
import calendar
import csv
import datetime
import math
import random
import sys
from pathlib import Path

import openpyxl

# ── 定数 ───────────────────────────────────────────────────────
DEFAULT_SUPERVISOR = "中原"

EXCLUDED_SHEETS: set = set()  # テンプレートファイルには不要なシートは含まれていない

MONTHLY_REQUIRED = {
    1:  {1: 91, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    2:  {1: 95, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    3:  {1: 95, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    4:  {1: 95, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    5:  {1: 94, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    6:  {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    7:  {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    8:  {1: 70, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    9:  {1: 70, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    10: {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    11: {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
    12: {1: 90, 2: 10, 3: 44, 4: 8, 5: 15, 6: 8},
}

NUM_TO_WORKS = {
    1: [
        ("掘削作業", ["マンホール布設", "汚水桝布設", "管布設", "水道掘削", "溝掘削", "法面掘削", "基礎掘削"]),
        ("土砂積込み作業", ["過積載防止", "周囲の確認", "積込み確認", "安全確認"]),
        ("走行操作作業", ["発進操作", "平坦地走行", "登坂操作", "降坂操作", "停止操作", "下車操作"]),
        ("毎日整備", ["目視点検", "グリース注入", "燃料補給", "清掃作業"]),
        ("始業前点検", ["目視点検", "備品確認", "始業確認", "安全点検"]),
    ],
    2: [
        ("作業開始前の安全装置等の点検作業", ["目視点検", "安全確認", "始業前確認", "安全装置確認"]),
        ("建設機械施工職種に必要な整理整頓作業", ["道具整理", "現場整理", "用具整頓", "工具整備"]),
        ("保護具の着用と服装の安全点検作業", ["保護具確認", "服装点検", "安全確認"]),
        ("雇入れ時等の安全衛生教育", ["安全教育", "衛生教育", "ルール確認"]),
    ],
    3: [
        ("掘削作業", ["手元作業", "補助作業", "埋戻し", "残土処理"]),
        ("締固め作業", ["埋戻し", "ランマ転圧", "転圧作業", "締固め確認"]),
        ("土工作業(対象職種・作業に係る手作業の作業）", ["手元作業", "蓋かさ上げ", "補助作業", "整地作業"]),
        ("積込み作業", ["手元作業", "補助積込み", "残土積込み"]),
        ("建設機械の管理及び点検・整備作業", ["バケット交換", "グリース注入", "清掃作業", "オイル点検"]),
    ],
    4: [
        ("安全衛生業務", ["荷物搬入", "現場清掃", "補助作業", "用具整備", "資材確認"]),
        ("建設機械施工職種に必要な整理整頓作業", ["道具整理", "現場整理", "工具整備"]),
    ],
    5: [
        ("建設機械の移送車両への積載及び移送作業", ["声出し誘導", "ユンボ回送", "機械移送", "積載補助", "誘導補助"]),
    ],
    6: [
        ("安全衛生業務", ["安全訓練", "倉庫整理", "現場内清掃", "安全管理", "安全確認"]),
    ],
}


# ── CSV 解析 ───────────────────────────────────────────────────

def parse_csv(csv_path: str) -> dict:
    """
    出勤簿CSVを読み込み、従業員ごとの出勤日情報を返す。

    CSV列: 日付,作業員コード,作業員名,現場名,所在地,出勤時刻,退勤時刻,実働時間,同行作業員,作業日報,種別

    出勤日の条件: 種別=="" かつ 出勤時刻 != ""

    Returns:
        {
            "EMP008": {
                "name": "アルフィアン ソレフル ハディ",
                "months": {
                    (2026, 5): [1, 2, 7, 8, ...]  # 出勤日リスト
                }
            },
            ...
        }
    """
    employees: dict = {}

    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code     = row["作業員コード"].strip()
            name     = row["作業員名"].strip()
            date_str = row["日付"].strip()
            clock_in = row["出勤時刻"].strip()
            leave    = row["種別"].strip()

            if not date_str:
                continue

            try:
                dt = datetime.datetime.strptime(date_str, "%Y/%m/%d")
            except ValueError:
                continue

            year, month, day = dt.year, dt.month, dt.day
            ym = (year, month)

            if code not in employees:
                employees[code] = {"name": name, "months": {}}
            if ym not in employees[code]["months"]:
                employees[code]["months"][ym] = set()

            # 出勤日: 種別が空かつ出勤時刻あり
            if leave == "" and clock_in:
                employees[code]["months"][ym].add(day)

    # set → sorted list
    for code in employees:
        employees[code]["months"] = {
            ym: sorted(days)
            for ym, days in employees[code]["months"].items()
        }

    return employees


# ── 番号割り振りロジック ─────────────────────────────────────────

def allocate_days(n_days: int, month: int) -> dict:
    reqs = MONTHLY_REQUIRED[month]
    min_d = {k: math.ceil(v / 8) for k, v in reqs.items()}
    total_min = sum(min_d.values())

    if n_days >= total_min:
        alloc = dict(min_d)
        alloc[1] += n_days - total_min
    elif n_days > 0:
        alloc = dict(min_d)
        deficit = total_min - n_days
        take_from_1 = min(deficit, alloc[1] - 1)
        alloc[1] -= take_from_1
        deficit -= take_from_1
        if deficit > 0:
            take_from_3 = min(deficit, alloc[3] - 1)
            alloc[3] -= take_from_3
    else:
        alloc = {k: 0 for k in reqs}

    return alloc


def shuffle_no_triple(lst: list, rng: random.Random) -> list:
    for _ in range(200):
        rng.shuffle(lst)
        if all(
            not (i >= 2 and lst[i] == lst[i - 1] == lst[i - 2])
            for i in range(len(lst))
        ):
            return lst
    return lst


# ── Excel 操作 ─────────────────────────────────────────────────

def find_template(wb) -> openpyxl.worksheet.worksheet.Worksheet:
    """テンプレートとなる既存の日誌シートを返す（最新月シートを優先）。"""
    candidates = [
        ws for ws in wb.worksheets
        if ws.title not in EXCLUDED_SHEETS
        and not ws.title.startswith("★")
        and not ws.title.startswith("■")
    ]
    if not candidates:
        raise ValueError("テンプレートになる日誌シートが見つかりません。")
    return candidates[0]


def update_header(ws, year: int, month: int, trainee_name: str):
    """ヘッダー部分（年月・氏名）だけ書き換える。"""
    ws["D5"] = f"（　　　{year}　　年　　{month}　　月分）"
    ws["E6"] = f"　　　　　　氏名　{trainee_name}"


def write_diary_rows(ws, year: int, month: int, working_days: list,
                     supervisor: str, rng: random.Random) -> dict:
    """行9〜39（日付別）に業務内容を書き込む。"""
    days_in_month = calendar.monthrange(year, month)[1]
    alloc = allocate_days(len(working_days), month)

    num_seq = []
    for num, count in alloc.items():
        num_seq.extend([num] * count)
    num_seq = shuffle_no_triple(num_seq, rng)

    work_idx = {k: 0 for k in NUM_TO_WORKS}

    for day in range(1, 32):
        row = 8 + day  # day1 → row9
        if day > days_in_month:
            for col in (1, 2, 4, 5, 8):
                ws.cell(row=row, column=col).value = None
            continue

        ws.cell(row=row, column=1).value = datetime.datetime(year, month, day)

        if day in working_days:
            seq_idx = working_days.index(day)
            num    = num_seq[seq_idx]
            works  = NUM_TO_WORKS[num]
            wi     = work_idx[num] % len(works)
            wname, shidou_list = works[wi]
            shidou = shidou_list[work_idx[num] % len(shidou_list)]
            work_idx[num] += 1

            ws.cell(row=row, column=2).value = wname
            ws.cell(row=row, column=4).value = num
            ws.cell(row=row, column=5).value = shidou
            ws.cell(row=row, column=8).value = supervisor
        else:
            ws.cell(row=row, column=2).value = "休み"
            ws.cell(row=row, column=4).value = None
            ws.cell(row=row, column=5).value = None
            ws.cell(row=row, column=8).value = None

    return alloc


def generate_sheet(wb, year: int, month: int, employee_code: str,
                   employee_name: str, working_days: list,
                   supervisor: str, rng: random.Random) -> dict:
    """1名分の日誌シートを生成または上書きする。"""
    short_name = employee_name.split()[0]  # 名前の最初のパート
    sheet_name = f"{year}.{month}_{short_name}"

    # 既存シートを削除して再生成
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    template = find_template(wb)
    new_ws = wb.copy_worksheet(template)
    new_ws.title = sheet_name
    wb.move_sheet(new_ws, offset=-len(wb.sheetnames) + 1)

    update_header(new_ws, year, month, employee_name)
    alloc = write_diary_rows(new_ws, year, month, working_days, supervisor, rng)

    return alloc


# ── サマリー表示 ───────────────────────────────────────────────

def print_summary(employee_name: str, month: int, working_days: list, alloc: dict):
    reqs = MONTHLY_REQUIRED[month]
    print(f"\n  【{employee_name}】  出勤: {len(working_days)}日  出勤日: {working_days}")
    for num in range(1, 7):
        req_h  = reqs[num]
        actual = alloc.get(num, 0) * 8
        ok     = "✓" if actual >= req_h else "△"
        print(f"    番号{num}: {actual:3d}h / {req_h:3d}h {ok}")


# ── メイン ─────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="出勤簿CSVから技能実習日誌を一括生成する"
    )
    parser.add_argument("csv", help="出勤簿CSVファイルのパス")
    parser.add_argument(
        "--excel", default=None,
        help="Excelファイルのパス（省略時: CSVと同じフォルダの *.xlsx を自動検出）"
    )
    parser.add_argument(
        "--trainees", default=None,
        help="処理する実習生の作業員コード（カンマ区切り）。省略時はCSV内の全従業員を処理。"
    )
    parser.add_argument("--supervisor", default=DEFAULT_SUPERVISOR, help="指導員名")
    parser.add_argument("--seed", type=int, default=None, help="乱数シード（再現性確保）")
    args = parser.parse_args()

    csv_path = Path(args.csv)
    if not csv_path.exists():
        print(f"エラー: CSVファイルが見つかりません: {csv_path}")
        sys.exit(1)

    # Excel ファイルを探す
    if args.excel:
        excel_path = Path(args.excel)
    else:
        candidates = list(csv_path.parent.glob("*.xlsx"))
        # ホームディレクトリも探す（日誌テンプレート.xlsx を優先）
        if not candidates:
            priority = list(Path.home().glob("日誌テンプレート.xlsx"))
            fallback = list(Path.home().glob("技能実習日誌*.xlsx"))
            candidates = priority + fallback
        if not candidates:
            print("エラー: Excelファイルが見つかりません。--excel で指定してください。")
            sys.exit(1)
        excel_path = candidates[0]
        print(f"Excel: {excel_path}")

    # CSV 読み込み
    print(f"CSV 読み込み中: {csv_path.name}")
    employees = parse_csv(str(csv_path))

    if not employees:
        print("エラー: CSVから出勤データが読み取れませんでした。")
        sys.exit(1)

    # 処理対象の絞り込み
    if args.trainees:
        target_codes = {c.strip() for c in args.trainees.split(",")}
        employees = {k: v for k, v in employees.items() if k in target_codes}
        if not employees:
            print(f"エラー: 指定したコード {target_codes} はCSVに存在しません。")
            sys.exit(1)

    print(f"\n処理対象: {len(employees)}名")
    for code, info in employees.items():
        total_months = sum(len(days) for days in info["months"].values())
        print(f"  {code}: {info['name']}  ({total_months}日分のデータ)")

    # Excel 生成
    rng = random.Random(args.seed)
    wb = openpyxl.load_workbook(str(excel_path))

    generated = []
    for code, info in employees.items():
        name = info["name"]
        for (year, month), working_days in sorted(info["months"].items()):
            if not working_days:
                print(f"  {name} {year}/{month}: 出勤日なし → スキップ")
                continue

            print(f"\n  [{name}] {year}年{month}月 — 出勤{len(working_days)}日")
            alloc = generate_sheet(
                wb, year, month, code, name,
                working_days, args.supervisor, rng
            )
            print_summary(name, month, working_days, alloc)
            generated.append(f"{year}.{month}_{name.split()[0]}")

    wb.save(str(excel_path))

    print(f"\n✓ 生成完了: {excel_path}")
    print(f"  作成シート: {', '.join(generated)}")
    print("\n次のステップ: Excelを開いて各シートを確認し、印刷してください。")


if __name__ == "__main__":
    main()
