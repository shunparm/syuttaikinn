#!/usr/bin/env python3
"""技能実習日誌 Excelからコンテンツデータベースを更新するスクリプト

使い方:
    python import_diary.py <Excelファイル>          # 取り込み
    python import_diary.py <Excelファイル> --dry-run # 確認のみ（書き込まない）
    python import_diary.py --show                   # 現在のDBを表示
"""

import argparse
import datetime
import json
import sys
from pathlib import Path

import openpyxl

DB_PATH = Path(__file__).parent / "training_content.json"


def load_db() -> dict:
    if DB_PATH.exists():
        with open(DB_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {
        "version": "1.0",
        "last_updated": "",
        "total_records": 0,
        "source_files": [],
        "entries": {str(n): {} for n in range(1, 7)},
    }


def save_db(db: dict) -> None:
    today = datetime.date.today().isoformat()
    db["last_updated"] = today
    with open(DB_PATH, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)


def is_diary_sheet(name: str) -> bool:
    """日誌シートかどうか判定（YYYY.M 形式）"""
    base = name.strip().split()[0].rstrip(".")
    parts = base.split(".")
    if len(parts) != 2:
        return False
    try:
        y, m = int(parts[0]), int(parts[1])
        return 2000 <= y <= 2100 and 1 <= m <= 12
    except ValueError:
        return False


def extract_from_workbook(wb: openpyxl.Workbook) -> tuple[list, list]:
    """ワークブックから出勤日誌レコードを抽出する。

    Returns:
        (records, skipped): recordsはdictのリスト、skippedは無視したシート名リスト
    """
    records = []
    skipped = []

    for sheet_name in wb.sheetnames:
        if not is_diary_sheet(sheet_name):
            skipped.append(sheet_name)
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=9, max_row=40, values_only=True):
            date_val  = row[0]
            work_name = row[1]
            num_raw   = row[3]
            guidance  = row[4]

            if not isinstance(date_val, datetime.datetime):
                continue
            if not work_name or work_name == "休み":
                continue
            if num_raw is None or guidance is None:
                continue
            try:
                num = int(num_raw)
            except (ValueError, TypeError):
                continue
            if num < 1 or num > 6:
                continue

            w = str(work_name).strip()
            g = str(guidance).strip()
            if w and g:
                records.append({"num": num, "work": w, "guidance": g})

    return records, skipped


def merge_records(db: dict, records: list) -> int:
    """レコードをDBにマージし、追加件数を返す。"""
    added = 0
    for r in records:
        key = str(r["num"])
        w, g = r["work"], r["guidance"]
        if w not in db["entries"][key]:
            db["entries"][key][w] = {}
        old = db["entries"][key][w].get(g, 0)
        db["entries"][key][w][g] = old + 1
        added += 1
    return added


def show_db() -> None:
    db = load_db()
    print(f"training_content.json")
    print(f"  最終更新: {db.get('last_updated', '不明')}")
    print(f"  総レコード: {db.get('total_records', 0)}件")
    print(f"  取り込み済みファイル: {len(db.get('source_files', []))}件")
    for f in db.get("source_files", []):
        print(f"    - {f}")
    print()
    for num in range(1, 7):
        key = str(num)
        entries = db["entries"].get(key, {})
        total = sum(sum(v.values()) for v in entries.values())
        print(f"【番号{num}】 {len(entries)}業務種 / {total}件")
        for w, gs in sorted(entries.items(), key=lambda x: -sum(x[1].values())):
            sub = sum(gs.values())
            tops = sorted(gs.items(), key=lambda x: -x[1])[:3]
            preview = "、".join(f"{g}({c})" for g, c in tops)
            print(f"  {w} ({sub}件): {preview}...")


def main() -> None:
    parser = argparse.ArgumentParser(description="技能実習日誌DBを更新する")
    parser.add_argument("excel", nargs="?", help="取り込むExcelファイルのパス")
    parser.add_argument("--dry-run", action="store_true", help="DBへの書き込みを行わない（確認のみ）")
    parser.add_argument("--show", action="store_true", help="現在のDBを表示して終了")
    args = parser.parse_args()

    if args.show:
        show_db()
        return

    if not args.excel:
        parser.print_help()
        sys.exit(1)

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"エラー: ファイルが見つかりません: {excel_path}")
        sys.exit(1)

    print(f"読み込み中: {excel_path.name}")
    wb = openpyxl.load_workbook(str(excel_path))
    records, skipped = extract_from_workbook(wb)

    print(f"  日誌シート: {len(wb.sheetnames) - len(skipped)}枚 / 無視: {len(skipped)}枚")
    if skipped:
        print(f"  無視したシート: {skipped}")
    print(f"  抽出レコード: {len(records)}件")

    if not records:
        print("取り込むデータがありませんでした。")
        return

    db = load_db()
    already = excel_path.name in db.get("source_files", [])
    if already:
        print(f"  ※ {excel_path.name} はすでに取り込み済みです（上書き追加します）")

    added = merge_records(db, records)
    db["total_records"] = db.get("total_records", 0) + added
    if excel_path.name not in db.get("source_files", []):
        db.setdefault("source_files", []).append(excel_path.name)

    if args.dry_run:
        print(f"\n[dry-run] 書き込みはしません（{added}件 追加予定）")
        return

    save_db(db)
    print(f"\n✓ 更新完了: {DB_PATH.name}  ({added}件 追加)")
    print()
    show_db()


if __name__ == "__main__":
    main()
